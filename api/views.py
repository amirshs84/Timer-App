from rest_framework import status, generics, views
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth.models import User
from django.utils import timezone
from django.db.models import Sum, Max, Count, Q, F
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from datetime import timedelta, datetime
from .models import UserProfile, Subject, StudySession, ConsultantTicket, School
from .serializers import (
    PhoneLoginSerializer,
    UserProfileSerializer,
    SubjectSerializer,
    StudySessionSerializer,
    ConsultantTicketSerializer,
    ManagerStudentListSerializer,
    ManagerDashboardKPISerializer,
    SchoolSerializer,
    AssignManagerSerializer
)
from .permissions import IsManager, IsSuperAdmin


import re
from django.contrib.auth.hashers import make_password, check_password


@api_view(['POST'])
@permission_classes([AllowAny])
def check_phone(request):
    """Check if phone number exists in database"""
    phone_number = request.data.get('phone_number')
    if not phone_number:
        return Response({'error': 'شماره تلفن الزامی است'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        user = User.objects.get(username=phone_number)
        profile = user.profile
        return Response({
            'exists': True,
            'has_password': profile.is_password_set
        })
    except User.DoesNotExist:
        return Response({
            'exists': False,
            'has_password': False
        })


def validate_password(password):
    """
    Validate password requirements:
    - At least 8 characters
    - Contains uppercase letter
    - Contains lowercase letter
    - Contains number
    - Contains special character
    """
    if len(password) < 8:
        return False, 'رمز عبور باید حداقل ۸ کاراکتر باشد'
    
    if not re.search(r'[A-Z]', password):
        return False, 'رمز عبور باید حداقل یک حرف بزرگ انگلیسی داشته باشد'
    
    if not re.search(r'[a-z]', password):
        return False, 'رمز عبور باید حداقل یک حرف کوچک انگلیسی داشته باشد'
    
    if not re.search(r'[0-9]', password):
        return False, 'رمز عبور باید حداقل یک عدد داشته باشد'
    
    if not re.search(r'[!@#$%^&*(),.?":{}|<>]', password):
        return False, 'رمز عبور باید حداقل یک علامت (!@#$%^&*...) داشته باشد'
    
    return True, ''


@api_view(['POST'])
@permission_classes([AllowAny])
def register_with_password(request):
    """Register new user with password"""
    phone_number = request.data.get('phone_number')
    password = request.data.get('password')
    password_confirm = request.data.get('password_confirm')
    
    if not phone_number or not password or not password_confirm:
        return Response({'error': 'همه فیلدها الزامی هستند'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if passwords match
    if password != password_confirm:
        return Response({'error': 'رمزهای عبور مطابقت ندارند'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Validate password
    is_valid, error_message = validate_password(password)
    if not is_valid:
        return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
    
    # Check if user already exists
    user_exists = User.objects.filter(username=phone_number).exists()
    
    if user_exists:
        user = User.objects.get(username=phone_number)
        # If user exists and has a password set, return error
        if hasattr(user, 'profile') and user.profile.is_password_set:
            return Response({'error': 'این شماره قبلاً ثبت شده است'}, status=status.HTTP_400_BAD_REQUEST)
        
        # If user exists but has no password (e.g. pre-created manager), update password
        user.set_password(password)
        user.save()
        
        profile = user.profile
        profile.is_password_set = True
        
        # If manager, skip profile completion
        if profile.role == 'manager':
            profile.is_profile_complete = True
            
        profile.save()
        
        # Generate tokens
        refresh = RefreshToken.for_user(user)
        
        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'profile': UserProfileSerializer(profile).data,
            'is_new_user': not profile.is_profile_complete
        }, status=status.HTTP_200_OK)
    
    # Create new user with password
    user = User.objects.create(
        username=phone_number,
        password=make_password(password)
    )
    
    profile = UserProfile.objects.create(
        user=user,
        phone_number=phone_number,
        is_password_set=True
    )
    
    # Generate tokens
    refresh = RefreshToken.for_user(user)
    
    return Response({
        'refresh': str(refresh),
        'access': str(refresh.access_token),
        'profile': UserProfileSerializer(profile).data,
        'is_new_user': True
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([AllowAny])
def login_with_password(request):
    """Login existing user with password"""
    phone_number = request.data.get('phone_number')
    password = request.data.get('password')
    
    if not phone_number or not password:
        return Response({'error': 'شماره تلفن و رمز عبور الزامی هستند'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        user = User.objects.get(username=phone_number)
        profile = user.profile
        
        # Check password
        if not check_password(password, user.password):
            return Response({'error': 'رمز عبور اشتباه است'}, status=status.HTTP_401_UNAUTHORIZED)
        
        # Generate tokens
        refresh = RefreshToken.for_user(user)
        
        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'profile': UserProfileSerializer(profile).data,
            'is_new_user': not profile.is_profile_complete
        }, status=status.HTTP_200_OK)
        
    except User.DoesNotExist:
        return Response({'error': 'کاربری با این شماره یافت نشد'}, status=status.HTTP_404_NOT_FOUND)


# Legacy OTP endpoints (keeping for backward compatibility)
@api_view(['POST'])
@permission_classes([AllowAny])
def request_otp(request):
    """
    Request OTP for phone number.
    In production, this would send SMS.
    For dev, it just returns success.
    """
    phone_number = request.data.get('phone_number')
    if not phone_number:
        return Response({'error': 'Phone number is required'}, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({'message': 'OTP sent', 'dev_code': '12345'})


@api_view(['POST'])
@permission_classes([AllowAny])
def phone_login(request):
    serializer = PhoneLoginSerializer(data=request.data)
    if serializer.is_valid():
        phone_number = serializer.validated_data['phone_number']
        otp = serializer.validated_data['otp']
        
        if otp != '12345':
            return Response(
                {'error': 'Invalid OTP'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user, created = User.objects.get_or_create(username=phone_number)
        profile, _ = UserProfile.objects.get_or_create(
            user=user,
            defaults={'phone_number': phone_number}
        )
        
        refresh = RefreshToken.for_user(user)
        
        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'profile': UserProfileSerializer(profile).data,
            'is_new_user': created or not profile.is_profile_complete
        }, status=status.HTTP_200_OK)
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UserProfileView(generics.RetrieveUpdateAPIView):
    """Get or update user profile"""
    serializer_class = UserProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return self.request.user.profile

    def perform_update(self, serializer):
        # Check if invitation_code is provided
        invitation_code = self.request.data.get('invitation_code')
        user_profile = self.get_object()
        
        if invitation_code:
            try:
                school = School.objects.get(invitation_code=invitation_code, is_active=True)
                serializer.save(is_profile_complete=True, school=school)
            except School.DoesNotExist:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({'invitation_code': 'کد دعوت نامعتبر است'})
        elif user_profile.school:
             # User already has a school, just update other fields
             serializer.save(is_profile_complete=True)
        else:
            # No code and no school
            from rest_framework.exceptions import ValidationError
            raise ValidationError({'invitation_code': 'کد دعوت الزامی است'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_study_status(request):
    """Update user's is_studying status when timer starts/stops"""
    is_studying = request.data.get('is_studying', False)
    profile = request.user.profile
    profile.is_studying = is_studying
    profile.save()
    return Response({'status': 'updated', 'is_studying': is_studying})


class SubjectListView(generics.ListCreateAPIView):
    """List user's subjects or create new one"""
    serializer_class = SubjectSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # Return only subjects created by this user
        return Subject.objects.filter(user=self.request.user)
    
    def perform_create(self, serializer):
        # Automatically assign the current user
        serializer.save(user=self.request.user)


class StudySessionListCreateView(generics.ListCreateAPIView):
    """List recent sessions or log a new one"""
    serializer_class = StudySessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return StudySession.objects.filter(user=self.request.user).order_by('-start_time')

    def perform_create(self, serializer):
        # Check if subject exists by name, or create it
        subject_data = self.request.data.get('subject_name')
        subject_color = self.request.data.get('subject_color', '#10b981')
        
        if subject_data:
            # Get or create subject for this specific user
            subject, _ = Subject.objects.get_or_create(
                name=subject_data,
                user=self.request.user,
                defaults={'color_code': subject_color}
            )
            serializer.save(user=self.request.user, subject=subject)
        else:
            # Fallback if subject ID provided
            serializer.save(user=self.request.user)


class DashboardStatsView(views.APIView):
    """Get aggregated stats for dashboard"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = now - timedelta(days=7)
        month_start = now - timedelta(days=30)

        # Helper to sum duration
        def get_duration(start_date):
            return StudySession.objects.filter(
                user=user,
                start_time__gte=start_date
            ).aggregate(total=Sum('duration_seconds'))['total'] or 0

        stats = {
            'today': get_duration(today_start),
            'week': get_duration(week_start),
            'month': get_duration(month_start),
            'total_sessions': StudySession.objects.filter(user=user).count()
        }
        
        return Response(stats)


class CreateTicketView(generics.CreateAPIView):
    serializer_class = ConsultantTicketSerializer
    permission_classes = [IsAuthenticated]
    
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


# ==================== Manager Panel Views ====================

class ManagerDashboardKPIView(views.APIView):
    """
    KPI Dashboard for managers - Bird's-eye view of school performance
    """
    permission_classes = [IsAuthenticated, IsManager]
    
    def get(self, request):
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        yesterday_start = today_start - timedelta(days=1)
        yesterday_end = today_start
        
        # Get all students in the manager's school
        manager_school = request.user.profile.school
        if not manager_school:
            return Response({
                'error': 'مدیر به هیچ مدرسه‌ای متصل نیست'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        students = User.objects.filter(
            profile__role='student',
            profile__school=manager_school
        )
        total_students = students.count()
        
        if total_students == 0:
            return Response({
                'avg_study_today': '0:00',
                'avg_study_today_seconds': 0,
                'change_percent': 0,
                'top_student': {'name': 'هیچ کس', 'total': 0},
                'absent_count': 0,
                'active_now': 0,
                'total_students': 0
            })
        
        # Average study time today
        today_totals = StudySession.objects.filter(
            user__profile__role='student',
            start_time__gte=today_start
        ).aggregate(total=Sum('duration_seconds'))['total'] or 0
        
        avg_today = today_totals // total_students if total_students > 0 else 0
        
        # Average study time yesterday for comparison
        yesterday_totals = StudySession.objects.filter(
            user__profile__role='student',
            start_time__gte=yesterday_start,
            start_time__lt=yesterday_end
        ).aggregate(total=Sum('duration_seconds'))['total'] or 0
        
        avg_yesterday = yesterday_totals // total_students if total_students > 0 else 0
        
        # Calculate change percent
        if avg_yesterday > 0:
            change_percent = ((avg_today - avg_yesterday) / avg_yesterday) * 100
        else:
            change_percent = 100 if avg_today > 0 else 0
        
        # Format avg_today as HH:MM
        hours = avg_today // 3600
        minutes = (avg_today % 3600) // 60
        avg_today_str = f"{hours}:{minutes:02d}"
        
        # Top student today
        top_student_data = StudySession.objects.filter(
            user__profile__role='student',
            start_time__gte=today_start
        ).values('user').annotate(
            total=Sum('duration_seconds')
        ).order_by('-total').first()
        
        if top_student_data:
            top_user = User.objects.get(id=top_student_data['user'])
            top_student = {
                'name': top_user.profile.full_name or top_user.profile.phone_number,
                'total': top_student_data['total']
            }
        else:
            top_student = {'name': 'هیچ کس', 'total': 0}
        
        # Absent students (no activity today)
        active_users_today = StudySession.objects.filter(
            start_time__gte=today_start
        ).values_list('user_id', flat=True).distinct()
        absent_count = total_students - len(set(active_users_today))
        
        # Active now (students with is_studying=True in their profile)
        active_now = UserProfile.objects.filter(
            role='student',
            school=manager_school,
            is_studying=True
        ).count()
        
        data = {
            'avg_study_today': avg_today_str,
            'avg_study_today_seconds': avg_today,
            'change_percent': round(change_percent, 1),
            'top_student': top_student,
            'absent_count': absent_count,
            'active_now': active_now,
            'total_students': total_students
        }
        
        serializer = ManagerDashboardKPISerializer(data)
        return Response(serializer.data)


class ManagerStudentListView(views.APIView):
    """
    List all students with their stats and trends
    """
    permission_classes = [IsAuthenticated, IsManager]
    
    def get(self, request):
        # Check if manager has a school
        manager_school = request.user.profile.school
        if not manager_school:
            return Response({
                'error': 'مدیر به هیچ مدرسه‌ای متصل نیست'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get filter parameters
        grade_filter = request.query_params.get('grade', None)
        olympiad_filter = request.query_params.get('olympiad', None)
        search = request.query_params.get('search', None)
        
        # Base queryset - only students in manager's school
        students = UserProfile.objects.filter(
            role='student',
            school=manager_school
        )
        
        # Apply filters
        if grade_filter:
            students = students.filter(grade=grade_filter)
        if olympiad_filter:
            students = students.filter(olympiad_field=olympiad_filter)
        if search:
            students = students.filter(
                Q(full_name__icontains=search) | Q(phone_number__icontains=search)
            )
        
        # Calculate stats for each student
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = now - timedelta(days=7)
        last_week_start = now - timedelta(days=14)
        
        students_data = []
        for profile in students:
            user = profile.user
            
            # Today's total
            today_total = StudySession.objects.filter(
                user=user,
                start_time__gte=today_start
            ).aggregate(total=Sum('duration_seconds'))['total'] or 0
            
            # This week's total
            week_total = StudySession.objects.filter(
                user=user,
                start_time__gte=week_start
            ).aggregate(total=Sum('duration_seconds'))['total'] or 0
            
            # Last week's total for trend
            last_week_total = StudySession.objects.filter(
                user=user,
                start_time__gte=last_week_start,
                start_time__lt=week_start
            ).aggregate(total=Sum('duration_seconds'))['total'] or 0
            
            # Calculate trend
            if last_week_total > 0:
                trend_percent = ((week_total - last_week_total) / last_week_total) * 100
                if trend_percent > 5:
                    trend = 'up'
                elif trend_percent < -5:
                    trend = 'down'
                else:
                    trend = 'stable'
            else:
                trend_percent = 100 if week_total > 0 else 0
                trend = 'up' if week_total > 0 else 'stable'
            
            # Last activity
            last_session = StudySession.objects.filter(user=user).order_by('-start_time').first()
            last_activity = last_session.start_time if last_session else None
            
            students_data.append({
                'user_id': user.id,
                'full_name': profile.full_name,
                'phone_number': profile.phone_number,
                'grade': profile.grade,
                'olympiad_field': profile.olympiad_field,
                'today_total': today_total,
                'week_total': week_total,
                'trend': trend,
                'trend_percent': round(trend_percent, 1),
                'last_activity': last_activity
            })
        
        # Sort by week_total descending
        students_data.sort(key=lambda x: x['week_total'], reverse=True)
        
        serializer = ManagerStudentListSerializer(students_data, many=True)
        return Response({
            'students': serializer.data,
            'count': len(students_data)
        })


class ManagerStudentProfileView(views.APIView):
    """
    Get detailed profile of a specific student (returns complete dashboard data)
    """
    permission_classes = [IsAuthenticated, IsManager]
    
    def get(self, request, user_id):
        # Ensure manager can only see students from their school
        try:
            manager_profile = request.user.profile
            manager_school = manager_profile.school
        except UserProfile.DoesNotExist:
            return Response({'error': 'Manager profile not found'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not manager_school:
            return Response({'error': 'Manager has no school assigned'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # First get the user to ensure existence
            user = User.objects.get(id=user_id)
            
            # Then check if they are a student in the manager's school
            # We use filter to avoid DoesNotExist if profile is missing (though unlikely)
            if not hasattr(user, 'profile'):
                 return Response({'error': 'User has no profile'}, status=status.HTTP_404_NOT_FOUND)
                 
            if user.profile.role != 'student':
                 return Response({'error': 'User is not a student'}, status=status.HTTP_404_NOT_FOUND)
            
            if user.profile.school != manager_school:
                 return Response({'error': 'Student not in your school'}, status=status.HTTP_403_FORBIDDEN)
                 
        except User.DoesNotExist:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Get time-based stats
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = now - timedelta(days=7)
        month_start = now - timedelta(days=30)
        
        def get_duration(start_date):
            return StudySession.objects.filter(
                user=user,
                start_time__gte=start_date
            ).aggregate(total=Sum('duration_seconds'))['total'] or 0
        
        # Get recent sessions (last 10)
        recent_sessions = StudySession.objects.filter(user=user).order_by('-start_time')[:10]
        sessions_data = []
        for session in recent_sessions:
            sessions_data.append({
                'id': session.id,
                'subject': session.subject.name if session.subject else 'بدون موضوع',
                'subject_color': session.subject.color_code if session.subject else '#gray',
                'duration_seconds': session.duration_seconds,
                'start_time': session.start_time.isoformat(),
                'end_time': session.end_time.isoformat() if session.end_time else None,
                'description': session.description or ''
            })
        
        # Get subject breakdown (total time per subject)
        subject_stats = StudySession.objects.filter(user=user).values(
            'subject__name', 'subject__color_code'
        ).annotate(
            total_seconds=Sum('duration_seconds')
        ).order_by('-total_seconds')
        
        subjects_breakdown = []
        for stat in subject_stats:
            if stat['subject__name']:
                subjects_breakdown.append({
                    'name': stat['subject__name'],
                    'color': stat['subject__color_code'] or '#10b981',
                    'total_seconds': stat['total_seconds']
                })
        
        # Get heatmap data (last 60 days)
        heatmap_start = now - timedelta(days=60)
        heatmap_sessions = StudySession.objects.filter(
            user=user,
            start_time__gte=heatmap_start
        ).values('start_time__date').annotate(
            total_seconds=Sum('duration_seconds')
        )
        
        heatmap_data = {}
        for item in heatmap_sessions:
            date_str = item['start_time__date'].isoformat()
            heatmap_data[date_str] = item['total_seconds']
        
        stats = {
            'student_name': user.profile.full_name,
            'phone_number': user.profile.phone_number,
            'grade': user.profile.grade,
            'olympiad_field': user.profile.olympiad_field,
            'today': get_duration(today_start),
            'week': get_duration(week_start),
            'month': get_duration(month_start),
            'total_sessions': StudySession.objects.filter(user=user).count(),
            'recent_sessions': sessions_data,
            'subjects_breakdown': subjects_breakdown,
            'heatmap_data': heatmap_data
        }
        
        return Response(stats)


class ManagerExportExcelView(views.APIView):
    """
    Export students performance data to Excel
    """
    permission_classes = [IsAuthenticated, IsManager]
    
    def get(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response(
                {'error': 'openpyxl library not installed. Install it with: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Get date range from query params
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            # Default: last 30 days
            end_date = timezone.now()
            start_date = end_date - timedelta(days=30)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "گزارش عملکرد"
        
        # Header style
        header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = ['نام', 'شماره تماس', 'پایه', 'رشته المپیاد', 'مجموع ساعات مطالعه', 'تعداد جلسات']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Get all students
        students = UserProfile.objects.filter(role='student')
        
        for profile in students:
            user = profile.user
            
            # Calculate total study time in date range
            total_seconds = StudySession.objects.filter(
                user=user,
                start_time__gte=start_date,
                start_time__lte=end_date
            ).aggregate(total=Sum('duration_seconds'))['total'] or 0
            
            hours = total_seconds / 3600
            
            session_count = StudySession.objects.filter(
                user=user,
                start_time__gte=start_date,
                start_time__lte=end_date
            ).count()
            
            ws.append([
                profile.full_name,
                profile.phone_number,
                profile.get_grade_display(),
                profile.get_olympiad_field_display(),
                f"{hours:.2f}",
                session_count
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=student_report_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
        
        wb.save(response)
        return response


class ManagerStudentReportPDFView(views.APIView):
    """
    Generate PDF report for a specific student
    """
    permission_classes = [IsAuthenticated, IsManager]
    
    def get(self, request, user_id):
        return Response(
            {'message': 'PDF generation will be implemented in next phase. For now, use Excel export.'},
            status=status.HTTP_501_NOT_IMPLEMENTED
        )


# ==================== SuperAdmin Panel Views ====================

class SuperAdminSchoolListCreateView(generics.ListCreateAPIView):
    """
    List all schools or create a new school (SuperAdmin only)
    """
    serializer_class = SchoolSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    queryset = School.objects.all()


class SuperAdminSchoolDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Get, update or delete a specific school (SuperAdmin only)
    """
    serializer_class = SchoolSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    queryset = School.objects.all()


class SuperAdminAssignManagerView(views.APIView):
    """
    Assign a manager to a school (SuperAdmin only)
    """
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    
    def post(self, request, school_id):
        try:
            school = School.objects.get(id=school_id)
        except School.DoesNotExist:
            return Response({'error': 'مدرسه یافت نشد'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = AssignManagerSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        phone_number = serializer.validated_data['phone_number']
        
        # Get or create user with this phone number
        user, created = User.objects.get_or_create(username=phone_number)
        profile, _ = UserProfile.objects.get_or_create(
            user=user,
            defaults={'phone_number': phone_number}
        )
        
        # Remove manager role from previous manager of this school (if exists)
        previous_managers = UserProfile.objects.filter(school=school, role='manager')
        for prev_manager in previous_managers:
            prev_manager.role = 'student'
            prev_manager.save()
        
        # Assign new manager
        profile.school = school
        profile.role = 'manager'
        profile.save()
        
        return Response({
            'message': f'مدیر با شماره {phone_number} به مدرسه {school.name} اختصاص یافت',
            'manager': {
                'phone_number': phone_number,
                'full_name': profile.full_name or 'تعیین نشده',
                'school': school.name
            }
        })


class SuperAdminSchoolMembersView(views.APIView):
    """
    Get list of all members (students + manager) of a school
    """
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    
    def get(self, request, school_id):
        try:
            school = School.objects.get(id=school_id)
        except School.DoesNotExist:
            return Response({'error': 'مدرسه یافت نشد'}, status=status.HTTP_404_NOT_FOUND)
        
        members = UserProfile.objects.filter(school=school)
        
        members_data = []
        for profile in members:
            members_data.append({
                'id': profile.user.id,
                'full_name': profile.full_name,
                'phone_number': profile.phone_number,
                'role': profile.get_role_display(),
                'grade': profile.get_grade_display() if profile.grade else '-',
                'is_profile_complete': profile.is_profile_complete
            })
        
        return Response({
            'school': SchoolSerializer(school).data,
            'members': members_data,
            'total_count': len(members_data)
        })

